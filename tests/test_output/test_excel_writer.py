"""ExcelWriter 테스트 모듈."""

import json
import os
import tempfile
from pathlib import Path
import pytest
from openpyxl import load_workbook
from src.output.excel_writer import ExcelWriter
from src.utils.exceptions import OutputError

class TestExcelWriter:
    """ExcelWriter 테스트 클래스."""
    
    def setup_method(self):
        """테스트 메서드 실행 전 초기화."""
        self.writer = ExcelWriter()
        self.temp_dir = tempfile.mkdtemp()
        self.test_file = Path(self.temp_dir) / "test_results.xlsx"
        
        # 테스트용 결과 데이터
        self.test_results = [{
            "ip": "192.168.1.10",
            "open_ports": [22, 80],
            "services": {22: "SSH", 80: "HTTP"},
            "os": "Linux",
            "checks": {
                "smbv1": False,
                "http_methods": ["GET", "POST"],
                "ftp_anonymous": True
            }
        }]
    
    def teardown_method(self):
        """테스트 메서드 실행 후 정리."""
        if self.test_file.exists():
            self.test_file.unlink()
        os.rmdir(self.temp_dir)
    
    def test_write_creates_file(self):
        """파일 생성 및 헤더 검증 테스트."""
        self.writer.write(self.test_results, str(self.test_file))
        
        assert self.test_file.exists()
        wb = load_workbook(self.test_file)
        ws = wb.active
        
        # 헤더 검증
        headers = [cell.value for cell in ws[1]]
        assert headers == self.writer.HEADERS
    
    def test_write_row_content(self):
        """데이터 행 내용 검증 테스트."""
        self.writer.write(self.test_results, str(self.test_file))
        
        wb = load_workbook(self.test_file)
        ws = wb.active
        
        # 데이터 행 검증
        row = list(ws.iter_rows(min_row=2, max_row=2))[0]
        assert row[0].value == self.test_results[0]["ip"]
        assert row[1].value == str(self.test_results[0]["open_ports"])
        assert row[2].value == json.dumps(self.test_results[0]["services"])
        assert row[3].value == self.test_results[0]["os"]
        assert row[4].value == str(self.test_results[0]["checks"]["smbv1"])
        assert row[5].value == str(self.test_results[0]["checks"]["http_methods"])
        assert row[6].value == str(self.test_results[0]["checks"]["ftp_anonymous"])
    
    def test_write_error(self):
        """저장 실패 시 예외 발생 테스트."""
        with pytest.raises(OutputError):
            self.writer.write(self.test_results, "/invalid/path/results.xlsx") 